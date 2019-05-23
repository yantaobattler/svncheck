# -*-coding:utf-8-*-
# !/usr/bin/env Python
# coding=utf-8

"""
##天津农商银行 源码清单检查工具
##闫涛@应用开发中心
##
##91请删除无用sheet页
##92建议项目名称填需求编号+需求名
##93请填写项目名称
##94请检查上线编号格式
##95上线编号不可为空
##01请删除空行
##02注意斜杠格式应该为/
##03路径应该以https开头
##04请确认源码名称及路径是否正确
##05版本号不存在
##06源码上传说明建议以软件需求编号开头
##07开发人员应于svn上传人用户名相同
##08提交日期不正确
##09源码清单不应包含文件夹记录
##10版本号不可为空
##11提交日期不可为空
##12开发人员不可为空
##13版本号须为数值型
##14日期须为日期型
"""


import svn
import datetime
import openpyxl
import svn.remote
import svn.local
from app_svn import tools
from app_svn.models import check_detail

authorpath = ''
write_file = ''
action = ''
result = ''
checking_temp = []


# 取SVN用户表查询中文名对应的英文用户名
# def checkauthor(name, author_ws):
#     author_list = []
#     author_ws_rows_len = len(list(author_ws.rows))
#     # 读每行
#     for row in range(3, author_ws_rows_len + 1):
#         if author_ws.cell(row=row, column=2).value and name == author_ws.cell(row=row, column=2).value.strip():
#             author_list.append(author_ws.cell(row=row, column=8).value.strip())
#             ##    print(name,'    ',author_list)
#
#     # 返回中文姓名对应的英文用户名list
#     return author_list


# 检查函数
def docheck(check_no, path):
    localpath = path
    # 打开检查文件
    wb = openpyxl.load_workbook(localpath)
    sheets = wb.get_sheet_names()

    # 检查每个sheet
    for sheetnum in range(len(sheets)):
        ws = wb.get_sheet_by_name(sheets[sheetnum])

        # 表头检查

        # 判断该sheet如果有源码则B3应为'程序源码清单'
        if ws['B1'].value or not ws['B3'].value or not '程序源码清单' == ws['B3'].value.strip():
            check_detail.objects.create(check_no=check_no, excel_line=3, line_name='head', problem='91')
            continue

            # 检查项目名称
        if ws['C5'].value:
            project_name = ws['C5'].value.strip()
            if not project_name[0:8].isdigit():
                check_detail.objects.create(check_no=check_no, excel_line=5, line_name='head', problem='92')
        else:
            check_detail.objects.create(check_no=check_no, excel_line=5, line_name='head', problem='93')

        # 检查上线编号
        if ws['C7'].value:
            update_no = ws['C7'].value.strip()
            if not (update_no[0:5] == 'UD-20' or update_no[0:5] == 'FD-20'):
                check_detail.objects.create(check_no=check_no, excel_line=7, line_name='head', problem='94')
        else:
            check_detail.objects.create(check_no=check_no, excel_line=7, line_name='head', problem='95')

        # 清单内容检查

        ws_rows_len = len(list(ws.rows))  # 行数

        # 循环读取每行明细
        for row in range(13, ws_rows_len + 1):
            passed = True
            # 判断该行是否为空
            if not ws.cell(row=row, column=2).value:
                check_detail.objects.create(check_no=check_no, excel_line=row, line_name='', problem='01')
                continue  # 该行为空直接下一行

            checking_name = ws.cell(row=row, column=2).value.strip()  # 源码名称
            checking_path = ws.cell(row=row, column=3).value.strip()  # 源码路径

            # 版本号是否为空
            if ws.cell(row=row, column=5).value:
                checking_revision = ws.cell(row=row, column=5).value
            else:
                check_detail.objects.create(check_no=check_no, excel_line=row, line_name=checking_name, problem='10')
                checking_revision = ''
                passed = False

            # 版本号应该是数值型
            if str(checking_revision).isdigit():
                pass
            else:
                check_detail.objects.create(check_no=check_no, excel_line=row, line_name=checking_name, problem='13')
                passed = False

            # 日期是否为空
            if ws.cell(row=row, column=6).value:
                checking_date = ws.cell(row=row, column=6).value
            else:
                checking_date = ''
                check_detail.objects.create(check_no=check_no, excel_line=row, line_name=checking_name, problem='11')
                passed = False

    #         # 日期应该是日期型
    #         try:
    #             if type(checking_date) == type('123'):
    #                 result = result + '14日期须为日期型'
    #                 ws.cell(row=row, column=9).value = result
    #                 continue
    #         except Exception as e:
    #             print(e)

            # 人员是否为空
            if ws.cell(row=row, column=7).value:
                checking_author = ws.cell(row=row, column=7).value.strip()
            else:
                check_detail.objects.create(check_no=check_no, excel_line=row, line_name=checking_name, problem='12')
                checking_author = ''
                passed = False

            # 路径格式整理

            # '\'替换为'/'
            if '\\' in checking_path:
                check_detail.objects.create(check_no=check_no, excel_line=row, line_name=checking_name, problem='02')
                passed = False

            # https开头
            if not checking_path.startswith(tools.svn.get('server')):
                check_detail.objects.create(check_no=check_no, excel_line=row, line_name=checking_name, problem='03')
                passed = False

            if not passed:
                continue

            # 拼url

            # 路径是否包含文件名
            if checking_path.endswith(checking_name):
                url = checking_path
            else:
                # 结尾有没有'/'
                if checking_path.endswith('/'):
                    url = checking_path + checking_name
                else:
                    url = checking_path + '/' + checking_name

            # 取svn服务器信息
            try:
                server_info = svn.remote.RemoteClient(url).info()
                if server_info.get('entry_kind') != 'file':  # 源码清单不应包含文件夹记录
                    check_detail.objects.create(check_no=check_no, excel_line=row, line_name=checking_name,
                                                problem='09')
                    continue

            except Exception:
                check_detail.objects.create(check_no=check_no, excel_line=row, line_name=checking_name,
                                            problem='04')
                continue

            # 检查
            # 是否最新版本
            # server_last_version = server_info.get('commit_revision')
            # if checking_revision == server_last_version:
            #     print('是最新版本')
            # else:
            #     print('不是最新版本')

            # 版本号是否存在
            try:
                server_log = svn.remote.RemoteClient(url).info(revision=checking_revision)
            except Exception:
                check_detail.objects.create(check_no=check_no, excel_line=row, line_name=checking_name,
                                            problem='05')
                continue

    #         # 检查message
    #         if not ws['I5'].value:
    #             if not log[-1].message.startswith(project_name[0:15]):
    #                 result = result + '06源码上传说明建议以软件需求编号开头'
    #         else:
    #             if not log[-1].message[0:8].isdigit():
    #                 result = result + '06源码上传说明建议以软件需求编号开头'

            # 检查人名
            server_log_name = server_log.get('commit_author')
            if server_log_name != checking_author:
                check_detail.objects.create(check_no=check_no, excel_line=row, line_name=checking_name,
                                            problem='07')

            # 检查日期
            server_log_date = server_log.get('commit_date')

            if type(checking_date) == datetime.datetime:  # 如果表里的是日期型，直接赋值为比较对象
                checking_datetime = checking_date
            elif type(checking_date) == str:  # 如果是字符型则格式化为日期型
                if '/' in checking_date:
                    date_l = checking_date.split('/')
                    checking_y = date_l[0]
                    checking_m = date_l[1]
                    checking_d = date_l[2]
                elif '-' in checking_date:
                    date_l = checking_date.split('-')
                    checking_y = date_l[0]
                    checking_m = date_l[1]
                    checking_d = date_l[2]
                elif len(checking_date) == 8:
                    checking_y = checking_date[0:4]
                    checking_m = checking_date[4:6]
                    checking_d = checking_date[6:8]
                else:
                    check_detail.objects.create(check_no=check_no, excel_line=row, line_name=checking_name,
                                                problem='08')
                    continue

                if checking_m.startswith('0') and len(checking_m) == 2:
                    checking_m = checking_m[-1]
                if checking_d.startswith('0') and len(checking_d) == 2:
                    checking_d = checking_d[-1]

                checking_datetime = datetime.datetime(checking_y, checking_m, checking_d)
            else:
                check_detail.objects.create(check_no=check_no, excel_line=row, line_name=checking_name,
                                            problem='08')
                continue

            if checking_datetime.strftime('%Y-%m-%d') != server_log_date.strftime('%Y-%m-%d'):
                check_detail.objects.create(check_no=check_no, excel_line=row, line_name=checking_name,
                                            problem='08')
